"""
Risco

Módulo que permite manipular arquivos .csv relacionados a relatórios de risco

Author: Juliano Kosloski - Automation Developer
Created: 23/08/2022 by Juliano Kosloski
"""

import openpyxl
import pyautogui
import os
import re
import datetime
import time

class Risco:
    
    def _getCurrentDate():
        """
        Gets and returns today's date
        """
        
        cd = datetime.date.today().strftime("%d%m%Y")
        return cd
       
    def getFiles() -> tuple:
        """
        Finds the paths to the downloaded files and returns a tuple with the file paths and the file name
        """
        
        cd = Risco._getCurrentDate() #gets current date
        date = cd[:2] + "_" + cd[2:4] + "_" + cd[4:] + "_" #puts the date in the right format for regex
        
        pattern = r"^psadeqewq_[{}].*[.csv]".format(date) #creates a regex to search the dir
        rx = re.compile(pattern) 

        for file in os.listdir(path = "C:/PATH/"): 
            
            if rx.match(file):
                
                file_name = file
                print("Um arquivo foi encontrado: " + file_name)
                
        pathRisco = "C:/PATH/"
        pathRiscoDetalhes = "C:/PATH/{}".format(file_name)
        
        return pathRisco, pathRiscoDetalhes, file_name
     
    def openFile(path : str) -> str:
        """
        Opens a file with the default Windows program, returning the file path
        
        params:
        path: a string with the file path
        """
        
        print("Opening file...")
        file = os.startfile(path) #opens the file for reading or writing
        time.sleep(10) #wait for the file to load
        
        return path
        
    def saveAsXLSX(path : str) -> str:
        """
        Navigates Excel GUI to save file as .xlsx, returning the new file path 
    
        params:
        path: a string with the file path
        """

        #this function solves the issue of losing data types when doing direct csv to xlsx conversion
        time.sleep(10)
        pyautogui.hotkey("alt", "a", interval = 2.0) #goes to file options in excel (PT-BR)
        time.sleep(10)
        pyautogui.press("a") #opens save as option
        time.sleep(30)
        pyautogui.hotkey("y", "4") #opens file type selection
        time.sleep(10)
        pyautogui.press("up", presses = 15) #navigates to .xlsx
        pyautogui.press("enter") #confirms the selection
        
        locPoint = pyautogui.locateCenterOnScreen('assets/save.png') #finds the save buttom and clicks it
        pyautogui.click(locPoint) 
        
        print("File saved as .xlsx...")
        newPath = path[:-4] + ".xlsx"
        time.sleep(20)
        return newPath
    
    def closeExcel() -> None:
        """
        Closes Excel
        """
        
        print("Closing Excel...")
        os.system("TASKKILL /F /IM EXCEL.exe")
        time.sleep(8)
        
    def prepFiles(riscoArq:str, riscoDet:str) -> tuple:
        """
        Prepares files by removing rows and columns, returning a tuple of the file paths
        
        params:
        riscoArq: string path to Arquivo.xlsx file
        riscoDet: string path to provisoes...xlsx file
        
        """
        
        wb = openpyxl.load_workbook(riscoDet) 
        ws = wb.worksheets[0]
        print("Deleting column G = 'Inad.' and Ativ. Prob.")
        ws.delete_cols(7, 2) #deletes column G, H
        wb.save(riscoDet)
        
        row_count = 0

        for i in range(ws.max_row + 1, 1, -1):   
            if str(ws.cell(row = i, column = 14).value) == "Recuperação de Prejuízo" or str(ws.cell(row = i, column = 14).value) == "Transferência para Prejuízo":
                    # ws.delete_rows(row[0].row)
                    ws.delete_rows(i, 1)
                    row_count += 1
            else:
                i+=1 #increases the counter only if no row is deleted
                
        print("Deleted {} Prejuízo rows...".format(row_count))          
        print("Saving prepared file...")
        wb.save(riscoDet)
        
        return riscoArq, riscoDet
    
    def clearMainFile(mainPath:str): 
        """
        Clears two sheets of the main file
        
        params:
        mainPath: str path to the main file
        """
        
        # opening the destination excel file 
        mainFile = mainPath
        main_wb = openpyxl.load_workbook(mainFile)
        main_ws = main_wb["Analítico - Associados"] #Analiticos Associados
        
        # calculate total number of rows and columns in source excel file
        maxr = main_ws.max_row
        maxc = main_ws.max_column
        
        print("Clearing Analitico-Associados...")
        for i in range (2, maxr + 1):
            for j in range (1, maxc + 1):
                # reading cell value from source excel file
                c = main_ws.cell(row = i, column = j)
                c.value = None
        
        #changing worksheet in main
        main_ws = main_wb["Resumo - Prévia Risco"] #Resumo P - Prévia Risco sheet
        
        print("Clearing Resumo - Prévia Risco...")
        for i in range (2, maxr + 1):
            for j in range (1, maxc + 1):
                # reading cell value from source excel file
                c = main_ws.cell(row = i, column = j)
                c.value = None
        
        print("Saving updated file...")
        main_wb.save(mainPath)
       
    def appendMainFile(path1 : str, path2 : str, mainPath:str) -> str:
        """
        Append data from the two prepared files into the main workbook and returns its path
        
        path1: path to the Arquivo.xlsx file 
        path2: path to the provisoes...xlsx file
        mainPath: path to the main file
        """
        
        # opening the source excel file
        arqFile = path1
        arq_wb = openpyxl.load_workbook(arqFile, data_only = True)
        arq_ws = arq_wb.worksheets[0]
        
        # opening another source excel file
        detFile = path2
        det_wb = openpyxl.load_workbook(detFile, data_only = True)
        det_ws = det_wb.worksheets[0]
        
        # opening the destination excel file 
        mainFile = mainPath #TODO change path
        main_wb = openpyxl.load_workbook(mainFile)
        main_ws = main_wb["Analítico - Associados"] #Analiticos Associados
        
        # calculate total number of rows and columns in source excel file
        maxr = arq_ws.max_row
        maxc = arq_ws.max_column
        
        print("Copying from Arquivo.xlsx to teste_risco.xlsx...")
        # copying the cell values from source excel file to destination excel file
        for i in range (1, maxr + 10000):
            for j in range (1, maxc + 1):
                # reading cell value from source excel file
                c = arq_ws.cell(row = i, column = j)
        
                # writing the read value to destination excel file
                main_ws.cell(row = i, column = j).value = c.value
        
        # calculate total number of rows and columns in another source excel file
        maxr = det_ws.max_row
        maxc = det_ws.max_column
        
        #changing worksheet in main
        main_ws = main_wb["Resumo - Prévia Risco"] #Resumo P - Prévia Risco sheet
        
        print("Copying from provisoes...xlsx to bd_risco_previa.xlsx...")
        
        # copying the cell values from another source excel file to destination excel file
        for i in range (1, maxr + 10000):
            for j in range (1, maxc + 1):
                # reading cell value from source excel file
                c = det_ws.cell(row = i, column = j)
        
                # writing the read value to destination excel file
                main_ws.cell(row = i, column = j).value = c.value
        
        # saving the destination excel file
        print("Saving updated file...")
        main_wb.save(str(mainFile))
        
        return mainFile
    
    def updateTables(mainFile:str) -> None:
        """
        Opens the main file in Excel to update the formula tables, then closes Excel
        
        params:
        mainFile: path to the main file
        """
        
        os.startfile(mainFile)
        time.sleep(15)
        pyautogui.hotkey('ctrl', 'b')
        time.sleep(15)
        os.system("TASKKILL /F /IM EXCEL.exe")
        time.sleep(5)
        
    def diffMotivo(mainFile:str):
        """
        Creates a table for the sum of the difference for each reason
        
        params:
        mainFile: path to the main file
        """
         
        #opens the main workbook with the formulas
        main_wb = openpyxl.load_workbook(mainFile)
        time.sleep(15)
        main_ws = main_wb["PRINCIPAIS MOTIVOS"]
        time.sleep(10)
        
        #opens a copy of the main workbook that only shows raw data
        data_wb = openpyxl.load_workbook(mainFile, data_only = True)
        time.sleep(15)
        data_ws = data_wb["PRINCIPAIS MOTIVOS"]
        time.sleep(10)
        print("Opening data-only workbook...")
        
        #gets the current date and adjusts format
        cd = Risco._getCurrentDate() 
        date = cd[:2] + "/" + cd[2:4] + "/" + cd[4:] 
        
        #stores the current date on the formula table
        c = main_ws.cell(row = 22, column = 3)
        c.value = date
        time.sleep(1)
        
        #iterates through columns until it finds the one corresponding to the current date

        r = 3
        col = 3
        
        print("Searching for the right column to copy data...")
        while True: 
            
            c = main_ws.cell(row = r, column = col)
            
            if c.value == date:
                print("A célula contém:" + str(c.value))
                break
                
            else:
                col += 1
        
        #gets data from the raw data_ws and stores in the main_ws
        data_row = 23
        data_col = 3
        main_row = 4
        main_col = col
        
        print("Copying raw data to PRINCIPAIS MOTIVOS...")
        while (data_row < 38):
            data = data_ws.cell(row = data_row, column = data_col).value
            main_ws.cell(row = main_row, column = main_col).value = data
            time.sleep(1)
            data_row += 1
            main_row += 1
        
        time.sleep(5)
        main_wb.save(str(mainFile)) 
        time.sleep(20)       
        print("Indicadores por Motivo calculados") 
             
    def diffAgencia(mainFile:str):
        """
        Creates a table for the sum of the difference for each agency
        
        params:
        mainFile: path to the main file
        """
        
        #opens the main workbook with the formulas
        main_wb = openpyxl.load_workbook(mainFile)
        time.sleep(15)
        main_ws = main_wb["Acompanhamento"]
        time.sleep(15)
        
        #opens a copy of the main workbook that only shows raw data
        data_wb = openpyxl.load_workbook(mainFile, data_only = True)
        time.sleep(15)
        data_ws = data_wb["Acompanhamento"]
        time.sleep(15)
        print("Opening data-only workbook...")
        
        #gets the current date and adjusts format
        cd = Risco._getCurrentDate() 
        date = cd[:2] + "/" + cd[2:4] + "/" + cd[4:] 
        
        #stores the current date on the formula table and saves
        c = main_ws.cell(row = 31, column = 4)
        c.value = date
        time.sleep(1)
        
        #iterates through columns until it finds the one corresponding to the current date
        r = 3
        col = 4

        print("Searching for the right column to copy data...")
        while True: 
            
            c = main_ws.cell(row = r, column = col)
            
            if c.value == date:
                print("A célula contém:" + str(c.value))
                break
                
            else:
                col += 1
                
        #gets data from the raw data_ws and stores in the main_ws
        data_row = 32
        data_col = 4
        main_row = 4
        main_col = col
        
        time.sleep(1)
        print("Copying raw data to Acompanhamento...")
        while (data_row < 57):
            data = data_ws.cell(row = data_row, column = data_col).value
            main_ws.cell(row = main_row, column = main_col).value = data
            time.sleep(1)
            data_row += 1
            main_row += 1
        
        time.sleep(5)
        main_wb.save(str(mainFile))
        time.sleep(20)
        print("Indicadores por Agencia calculados")
       
    def saveFiles(mainFile:str) -> None:
        """
        Loads main workbook and saves it to two different folders: one to update the dashboards and another as backup
        
        params:
        mainFile: path to the main file
        """
        
        #gets the current date and adjusts format
        cd = Risco._getCurrentDate() 
        
        #builds the two paths
        bi_path = "C:/PATH/Bweq_ewqq.xlsx"
        backup_path = "C:/PATH//qwqqwAeq{}.xlsx".format(cd)
        
        #opens the main workbook and saves it to the BI folder
        main_wb = openpyxl.load_workbook(mainFile)
        time.sleep(15)
        main_wb.save(bi_path)
        
        time.sleep(15)
        #opens the main workbook and saves it to the backup folder
        main_wb = openpyxl.load_workbook(mainFile)
        time.sleep(15)
        main_wb.save(backup_path)
        time.sleep(10)
        
        print("Documentos salvos nas respectivas pastas. Encerrando programa.")
    
    

    